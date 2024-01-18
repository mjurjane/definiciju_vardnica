from openpyxl import Workbook, load_workbook
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

import time

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

def interneta_meklesana(vards):
    url = "https://tezaurs.lv"
    driver.get(url)

    find = driver.find_element(By.ID, "searchField")
    find.send_keys(vards)

    find = driver.find_element(By.CLASS_NAME, "form-search-submit")
    find.click()
    time.sleep(2)

    try:
        find = driver.find_element(By.CLASS_NAME, "dict_Gloss")
        return find.text

    except NoSuchElementException:
        return "Tezaurā šim vārdam skaidrojuma nav"

    # find = driver.find_element(By.CLASS_NAME, "dict_Gloss")
    # if find is None:
    #     return "Tezaurā šim vārdam skaidrojuma nav"
    # else:
    #     return find.text

wb = load_workbook('vardnica.xlsx')
ws = wb.active
max_row = ws.max_row

for row in range (2,max_row+1):
    a = str(ws['A' + str(row)].value)
    skaidrojums = interneta_meklesana(a)
    ws['B' + str(row)].value = skaidrojums
    print(skaidrojums)

wb.save("vardnica_ar_skaidrojumiem.xlsx")
wb.close()